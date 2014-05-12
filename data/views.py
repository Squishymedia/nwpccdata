from django.http import HttpResponse, Http404, HttpResponseForbidden
from django.shortcuts import render_to_response
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.csrf import csrf_exempt
from models import *
from rest_framework import viewsets, generics
from rest_framework.response import Response
from rest_framework.decorators import action
from data.serializers import DataSetSerializer, ShortDataSetSerializer, DataSetRevisionSerializer, ShortDataSetRevisionSerializer
from openpyxl import load_workbook
import datetime
import json

class DataSetViewSet(viewsets.ModelViewSet):
    model = DataSet

    def list(self, request):
        queryset = DataSet.objects.all()
        serializer = ShortDataSetSerializer(queryset, many=True)
        return Response(serializer.data)

    def detail(self, request, id=None):
        try:
            queryset = DataSet.objects.get(id=id)
        except ObjectDoesNotExist:
            raise Http404
        serializer = DataSetSerializer(queryset)
        return Response(serializer.data)

class DataSetRevisionViewSet(viewsets.ModelViewSet):
    model = DataSetRevision

    def detail(self, request, id=None, revision=None):
        try:
            queryset = DataSetRevision.objects.get(dataset__id=id, revision_number=revision)
        except ObjectDoesNotExist:
            raise Http404
        serializer = DataSetRevisionSerializer(queryset)
        data = serializer.data
        data["data"] = json.loads(queryset.data)
        data["column_names"] = json.loads(queryset.column_names)
        return Response(data)

    def current(self, request, id=None, revision=None):
        try:
            queryset = DataSet.objects.get(id=id).current_revision
        except ObjectDoesNotExist:
            raise Http404
        serializer = DataSetRevisionSerializer(queryset)
        data = serializer.data
        data["data"] = json.loads(queryset.data)
        data["column_names"] = json.loads(queryset.column_names)
        return Response(data)

    def latest(self, request, id=None):
        try:
            queryset = DataSetRevision.objects.filter(dataset__id=id).order_by('-revision_number')[0]
        except ObjectDoesNotExist:
            raise Http404
        serializer = DataSetRevisionSerializer(queryset)
        data = serializer.data
        data["data"] = json.loads(queryset.data)
        data["column_names"] = json.loads(queryset.column_names)
        return Response(data)

@csrf_exempt
def approve_set_data(request, id, revision):
    if not (request.method == 'POST'):
        return HttpResponseForbidden('{"error": "GET not allowed, POST only."}')
    try:
        data_revision = DataSetRevision.objects.get(dataset__id=id, id=revision)
    except ObjectDoesNotExist:
        raise Http404
    data_revision.dataset.current_revision = data_revision
    data_revision.save()
    return HttpResponse('{"message": "Data set ' + str(id) + ' is now using revision ' + str(revision) + '."}', status=200)

@csrf_exempt
def post_new_xls(request, id):
    if not (request.method == 'POST'):
        return HttpResponseForbidden('{"error": "GET not allowed, POST only."}')
    if request.method == 'POST':
        sheet_name = str(request.POST.get('sheet', default='Projects'))
        header_row = int(request.POST.get('header_row', default=1))
        comment = request.POST.get('revision_comment', default='')
        current = request.POST.get('make_canonical', default=False)
        # There is no security in this application. That will needed to be added if future revisions are created.
        user = request.POST.get('user', default='Anonymous')
        print 'Sheet Name: ' + str(sheet_name) + ', Header Row: ' + str(header_row) + ', User: ' + user + ', Comment: ' + comment + ', Current: ' + str(current)
        print 'Files: '
        print str(request.FILES)
        file = request.FILES['file']
        filedata = file.read()
        # This won't support concurrency. If this application goes forward, we'll need to create a random
        # temp file and delete it when we're done.
        f = open('/tmp/tempfile.tmp', 'wb')
        f.write(filedata)
        f.close()
        sheet = load_workbook(filename=r'/tmp/tempfile.tmp')
        sheet_ranges = sheet[sheet_name]
        sheet_names = sheet.get_sheet_names()
        print 'Workbook: ' + str(sheet)
        print 'Projects Sheet: ' + str(sheet_ranges)        
        print 'Named Ranges: ' + str(sheet.get_named_ranges())
        print 'Sheet Names: ' + str(sheet.get_sheet_names())
        #print 'Sheet 2: ' + sheet_names[1]
        #print 'Get sheet 2: ' + str(sheet[sheet_names[1]])
        print 'Sheet dimensions: ' + str(sheet_ranges.calculate_dimension())
        columns = sheet_ranges.get_highest_column()
        print 'Highest column: ' + str(columns)
        rows = sheet_ranges.get_highest_row()
        print 'Highest row: ' + str(rows)
        rowdata = []
        for y in range (0, rows):
            thisrow = []
            for x in range (0, columns):
                thisrow.append( sheet_ranges.cell(column=x, row=y).value )
            rowdata.append(thisrow)
        revision = DataSetRevision()
        revision.dataset = DataSet.objects.get(id=id)
        revisions = DataSetRevision.objects.filter(dataset=revision.dataset).order_by('-revision_number')
        if revisions.count() > 0:
            revision.revision_number = revisions[0].revision_number + 1
        else:
            revision.revision_number = 1
        revision.user = 'Administrator'
        revision.comment = comment
        if header_row:
            revision.column_names = json.dumps(rowdata[header_row-1], default=date_default)
            revision.data = json.dumps(rowdata[header_row:], default=date_default)
        else:
            revision.data = json.dumps(rowdata, default=date_default)
        print 'Stored Data Size: ' + str(len(revision.data)) + ' bytes'
        revision.save()
        message = 'A new data set has been created for set ' + str(id) + ' as revision ' + str(revision.revision_number) + '.'
        if current:
            revision.dataset.current_revision = revision
            revision.dataset.save()
            message += ' This data set has been made the new canonical data set.'
        return HttpResponse('{"message":' + message + '}', status=201) # Created

def date_default(obj):
    """JSON Serializer. Required because the Python serializer doesn't know how to handle a DateTime object."""

    if isinstance(obj, datetime.datetime):
        return str(obj)
    else:
        return JSONEncoder.default(self, obj)

def tests(self):
    return render_to_response('tests.htm')

def index(self):
    return render_to_response('index.htm')
